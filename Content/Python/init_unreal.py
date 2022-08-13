# Copyright 2022 Proletariat, Inc.

import unreal
import os
import io
import traceback
from pathlib import Path
import openpyxl
import json
import pm_xlsx_field_parser as fp
import pm_xlsx_parser as xlsx_parser
from importlib import *

reload(xlsx_parser)
reload(fp)


@unreal.uclass()
class PMXlsxImporterPythonBridgeImpl(unreal.PMXlsxImporterPythonBridge):

    @unreal.ufunction(override=True)
    def read_worksheet_names(self, absolute_file_path):
        unreal.log("Reading xlsx file \"{0}\"".format(absolute_file_path))

        with open(absolute_file_path, "rb") as f:
            in_mem_file = io.BytesIO(f.read())
        workbook = openpyxl.load_workbook(in_mem_file, read_only=True, data_only=True)
        unreal.log("All sheet names: {0}".format(workbook.sheetnames))
        return workbook.sheetnames

    @unreal.ufunction(override=True)
    def read_worksheet_asset_names(self, absolute_file_path, worksheet_name, header_row, data_start_row):
        result = unreal.PMXlsxImporterPythonBridgeAssetNames()
        try:
            with open(absolute_file_path, "rb") as f:
                in_mem_file = io.BytesIO(f.read())
            workbook = openpyxl.load_workbook(in_mem_file, read_only=True, data_only=True)
            worksheet = workbook[worksheet_name]

            asset_names = []

            for row in worksheet.iter_rows(min_row=data_start_row):
                name_cell = row[0]  # Name should always be the first column
                if not name_cell.value or name_cell.value.isspace():
                    break  # not valid since this row
                asset_names.append(str(name_cell.value))

            result.asset_names = asset_names
        except (ValueError, Exception) as ex:
            result.error = traceback.format_exc()
        finally:
            return result

    @unreal.ufunction(override=True)
    def read_worksheet_as_json(self, absolute_file_path, worksheet_name, header_row, data_start_row,
                               worksheet_type_info):
        result = unreal.PMXlsxImporterPythonBridgeJsonString()
        try:
            parser = xlsx_parser.PMXlsxParser(absolute_file_path, worksheet_name, worksheet_type_info)

            # parse header row
            parser.parse_header_row(header_row)

            # parse data rows
            data_list = parser.parse_data(data_start_row)

            # convert data to json string
            json_string = json.dumps(data_list, indent=4)

            # write json string to Intermediate folder for debug purpose
            json_parent_dir = Path(absolute_file_path).stem
            json_file_name = worksheet_name + '.json'
            json_output_file = Path(os.path.join(
                unreal.Paths.convert_relative_path_to_full(unreal.Paths.project_intermediate_dir()), 'XlsxJsonFiles',
                json_parent_dir, json_file_name))
            json_output_file.parent.mkdir(exist_ok=True, parents=True)
            json_output_file.write_text(json_string)

            result.json_string = json_string

        except fp.ValidationError as ex:
            result.error = str(ex)
        except Exception as ex:
            # handle all other exceptions
            result.error = traceback.format_exc()
        finally:
            return result
